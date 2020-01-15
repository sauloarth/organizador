from openpyxl import load_workbook
from openpyxl import Workbook

def lotacaoswitcher(cod):
	switcher = {
		"390600000000": "390600000000 - DIRAPS",
		"390601000000"	: "390601000000 - NVEPI/DIRAPS",
		"390602000000"	: "390602000000 - GPMA/DIRAPS",
		"390602010000"	: "390602010000 - NCAIS/DIRAPS",
		"390602020000"	: "390602020000 - NGC/DIRAPS",
		"390604000000"	: "390604000000 - GAPAPS/DIRAPS",
		"390605000000"	: "390605000000 - GENF/DIRAPS",
		"390606000000"	: "390606000000 - GERNO",
		"390607000000"	: "390607000000 - GSAP1-PLA",
		"390607010000"	: "390607010000 - GSAP1-PLA/UBS1-PLA",
		"390608000000"	: "390608000000 - GSAP2-PLA",
		"390608010000"	: "390608010000 - GSAP2-PLA/UBS2-PLA",
		"390608020000"	: "390608020000 - GSAP2-PLA/UBS7-PLA",
		"390609000000"	: "390609000000 - GSAP3-PLA",
		"390609010000"	: "390609010000 - GSAP3-PLA/UBS3-PLA",
		"390609020000"	: "390609020000 - GSAP3-PLA/UBS18-PLA",
		"390609030000"	: "390609030000 - GSAP3-PLA/UBS19-PLA",
		"390610000000"	: "390610000000 - GSAP4-PLA",
		"390610010000"	: "390610010000 - GSAP4-PLA/UBS10-PLA",
		"390610020000"	: "390610020000 - GSAP4-PLA/UBS16-PLA",
		"390610030000"	: "390610030000 - GSAP4-PLA/UBS17-PLA",
		"390611000000"	: "390611000000 - GSAP5-PLA",
		"390611010000"	: "390611010000 - GSAP5-PLA/UBS13-PLA",
		"390611020000"	: "390611020000 - GSAP5-PLA/UBS14-PLA",
		"390611030000"	: "390611030000 - GSAP5-PLA/UBS15-PLA",
		"390612000000"	: "390612000000 - GSAP6-PLA",
		"390612010000"	: "390612010000 - GSAP6-PLA/UBS8-PLA",
		"390612020000"	: "390612020000 - GSAP6-PLA/UBS9-PLA",
		"390613000000"	: "390613000000 - GSAP7-PLA",
		"390613010000"	: "390613010000 - GSAP7-PLA/UBS11-PLA",
		"390613020000"	: "390613020000 - GSAP7-PLA/UBS12-PLA",
		"390614000000"	: "390614000000 - GSAP8-PLA",
		"390614010000"	: "390614010000 - GSAP8-PLA/UBS4-PLA",
		"390615000000"	: "390615000000 - GSAP9-PLA",
		"390615010000"	: "390615010000 - GSAP9-PLA/UBS5-PLA",
		"390615020000"	: "390615020000 - GSAP9-PLA/UBS6-PLA",
		"390616000000"	: "390616000000 - GSAP1-SOB",
		"390616010000"	: "390616010000 - GSAP1-SOB/UBS1-SOB",
		"390616020000"	: "390616020000 - GSAP1-SOB/UBS5-SOB",
		"390616030000"	: "390616030000 - GSAP1-SOB/USB6-SOB",
		"390617000000"	: "390617000000 - GSAP2-SOB",
		"390617010000"	: "390617010000 - GSAP2-SOB/UBS2-SOB",
		"390618000000"	: "390618000000 - GSAP3-SOB",
		"390618010000"	: "390618010000 - GSAP3-SOB/UBS1-SOB2",
		"390619000000"	: "390619000000 - GSAP4-SOB",
		"390619010000"	: "390619010000 - GSAP4-SOB/UBS3-SOB",
		"390619020000"	: "390619020000 - GSAP4-SOB/UBS4-SOB",
		"390620000000"	: "390620000000 - GSAP5-SOB",
		"390620100000"	: "390620100000 - GSAP5-SOB/UBS2-SOB2",
		"390621000000"	: "390621000000 - GSAP6-SOB",
		"390621010000"	: "390621010000 - GSAP6-SOB/UBS3-SOB2",
		"390621020000"	: "390621020000 - GSAP6-SOB/UBS4-SOB2",
		"390621020000"	: "390621020000 - GSAP6-SOB/UBS5-SOB2",
		"390621030000"	: "390621030000 - GSAP6-SOB/UBS6-SOB2",
		"390622000000"	: "390622000000 - GSAP7-SOB",
		"390622010000"	: "390622010000 - GSAP7-SOB/UBS1-FERCAL",
		"390622020000"	: "390622020000 - GSAP7-SOB/UBS2-FERCAL",
		"390623000000"	: "390623000000 - CERPIS",
		"390623010000"	: "390623010000 - FARMÁCIA/CERPIS",
		"390624000000"	: "390624000000 - GEQUALI/DIRAPS"
	}
	return switcher.get(cod, "")

wb = Workbook();
ws = wb.create_sheet("Controle", 0)

controle = load_workbook('controle.xlsx')

cx=1
cy=1

allworksheets = controle.sheetnames

for i, name in enumerate(allworksheets):
	gsap = controle[name]
	
	gsapx=5
	gsapy=3

	while gsap.cell(row=gsapx, column=3).value != ("" or "None"):
		
	
		for mes in range(2,14):
			referencia = gsap.cell(row=gsapx, column=(mes * 2))
			if referencia.value == 'C':
				
				lotacao = lotacaoswitcher(gsap.cell(row=gsapx, column=2).value)

				ws.cell(row=cx, column=cy).value = lotacao

				ws.cell(row=cx, column=cy+1).value = gsap.cell(row=3, column=(mes * 2)).value
				ws.cell(row=cx, column=cy+2).value = '2019'
				servidor = str(gsap.cell(row=gsapx, column=1).value).zfill(8) + " - " + gsap.cell(row=gsapx, column=3).value
				ws.cell(row=cx, column=cy+3).value = servidor
				ws.cell(row=cx, column=cy+4).value = 'Conferido'
				observacao = ""
				if gsap.cell(row=gsapx, column=(mes * 2) + 1).value != ("" or "None"):
					observacao = gsap.cell(row=gsapx, column=(mes * 2) + 1).value

				ws.cell(row=cx, column=cy+5).value = observacao


			if referencia.value == 'D':
				
				lotacao = lotacaoswitcher(gsap.cell(row=gsapx, column=2).value)

				ws.cell(row=cx, column=cy).value = lotacao

				ws.cell(row=cx, column=cy+1).value = gsap.cell(row=3, column=(mes * 2)).value
				ws.cell(row=cx, column=cy+2).value = '2019'
				servidor = str(gsap.cell(row=gsapx, column=1).value).zfill(8) + " - " + gsap.cell(row=gsapx, column=3).value
				ws.cell(row=cx, column=cy+3).value = servidor
				ws.cell(row=cx, column=cy+4).value = 'Devolvido'
				observacao = ""
				if gsap.cell(row=gsapx, column=(mes * 2) + 1).value != ("" or "None"):
					observacao = gsap.cell(row=gsapx, column=(mes * 2) + 1).value

				ws.cell(row=cx, column=cy+5).value = observacao

			gsapx += 1
			wb.save('novocontrole.xlsx')
